import streamlit as st
import pandas as pd
import numpy as np
from pandas.api.types import is_numeric_dtype
# from utils.queries import *
from utils.user import *
import openpyxl
import os
from st_aggrid import GridUpdateMode, JsCode, StAggridTheme
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode, GridUpdateMode
from st_aggrid.shared import StAggridTheme
import streamlit.components.v1 as components
from rapidfuzz import fuzz
import re
from utils.constants.general_constants import *
import io


# Formata valores numéricos e datas 
def filtra_formata_df(df, coluna_data, id_casa, start_date, end_date):
    if id_casa != 157: 
        df_filtrado = df[df['ID_Casa'] == id_casa] 
        df_filtrado = df_filtrado[(df_filtrado[coluna_data] >= start_date) & (df_filtrado[coluna_data] <= end_date)] 
    else:    
        df_filtrado = df[(df[coluna_data] >= start_date) & (df[coluna_data] <= end_date)] 

    # Copia para formatação brasileira de colunas numéricas 
    df_formatado = df_filtrado.copy() 
    
    # Aplica formatação brasileira em colunas numéricas 
    for col in df_formatado.select_dtypes(include='object').columns: 
        if col != "Doc_NF":
            df_formatado[col] = df_formatado[col].apply(format_brazilian) 
    
    # Aplica formatação brasileira em colunas de data 
    for col in df_formatado.select_dtypes(include='datetime').columns: 
        df_formatado[col] = pd.to_datetime(df_formatado[col]).dt.strftime('%d-%m-%Y %H:%M') 
    return df_filtrado, df_formatado

# Recebe df filtrado e só formata campos numéricos e de data
def formata_df(df):
    # Copia para formatação brasileira de colunas numéricas 
    df_formatado = df.copy() 
    
    # Aplica formatação brasileira em colunas numéricas 
    for col in df_formatado.select_dtypes(include=['object', 'number']).columns: 
        if col != "Doc_NF" and "ID" not in col:
            df_formatado[col] = df_formatado[col].apply(format_brazilian) 
    
    # Aplica formatação brasileira em colunas de data 
    for col in df_formatado.select_dtypes(include='datetime').columns: 
        df_formatado[col] = pd.to_datetime(df_formatado[col]).dt.strftime('%d-%m-%Y %H:%M') 
    return df_formatado


# Funções para colorir df de acordo com condições e exibir legenda 
def colorir_conciliacao(row):
    if row['Conciliação'] != '0,00':
        return ['background-color: #e6937e; color: black;'] * len(row)
    else:
        return [''] * len(row)  # Conciliados não pinta


def colorir_linhas(df, coluna_duplicados, coluna_doc, coluna_aprov, principal):
    def aplicar_linha(row):
        estilos = [''] * len(row)

        if principal == 'despesa':
            # prioridade: extrato ausente > duplicado > não aprovado
            if pd.isna(row['ID_Extrato_Bancario']): # vermelho
                estilos = ['background-color: #e6937e; color: black;'] * len(row)

            elif df[coluna_duplicados].value_counts().get(row[coluna_duplicados], 0) > 1: # amarelo
                estilos = ['background-color: #ffffae; color: black;'] * len(row)

            elif df['ID_Extrato_Bancario'].value_counts().get(row['ID_Extrato_Bancario'], 0) > 1: # cinza
                estilos = ['background-color: #bfbfbf; color: black;'] * len(row)

            # só aplica laranja se não cair nas regras acima
            if (
                all(s == '' for s in estilos) and
                coluna_duplicados not in ['ID_Bloqueio', 'Mutuo_ID'] and
                (pd.isna(row[coluna_doc]) or pd.isna(row[coluna_aprov]))
            ):
                estilos = ['background-color: #ffac34; color: black;'] * len(row)

        elif principal == 'extrato':
           # prioridade: despesa ausente > duplicado > não aprovado
            if pd.isna(row['ID_Despesa']): # vermelho
                estilos = ['background-color: #e6937e; color: black;'] * len(row)

            elif df['ID_Despesa'].value_counts().get(row['ID_Despesa'], 0) > 1: # amarelo
                estilos = ['background-color: #ffffae; color: black;'] * len(row)

            elif df['ID_Extrato_Bancario'].value_counts().get(row['ID_Extrato_Bancario'], 0) > 1: # cinza
                estilos = ['background-color: #bfbfbf; color: black;'] * len(row)

        return estilos
    return aplicar_linha


# Exibe legenda para linhas pintadas do dataframe
def exibir_legenda(parametro): 
    if parametro == 'conciliacao':
        span = 'Dias não conciliados'
        st.markdown(
            f"""
            <div style="display: flex; align-items: center; padding:10px; border:1px solid #ccc; border-radius:8px";>
                <div style="width: 15px; height: 15px; background-color: #e6937e; border: 1px solid #ccc; margin-right: 10px;"></div>
                <span style="font-size: 14px">{span}</span>
            </div>
            """,
            unsafe_allow_html=True
        )

    elif parametro == 'contas':
       st.markdown(
            f"""
            <div style="display: flex; align-items: center; padding:10px; border:1px solid #ccc; border-radius:8px;">
                <div style="width: 15px; height: 15px; background-color: #e6937e; border: 1px solid #ccc; margin-right: 8px;"></div>
                <span style="margin-right: 15px; font-size: 14px;">Despesa não encontrada no extrato bancário</span>
                <div style="width: 15px; height: 15px; background-color: #ffffae; border: 1px solid #ccc; margin-right: 8px;"></div>
                <span style="margin-right: 15px; font-size: 14px;">Mesma despesa mapeada com mais de um item no extrato bancário (correspondência incorreta)</span>
                <div style="width: 15px; height: 15px; background-color: #bfbfbf; border: 1px solid #ccc; margin-right: 8px;"></div>
                <span style="margin-right: 15px; font-size: 14px;">Despesas diferentes mapeadas com um mesmo item no extrato bancário (correspondência incorreta)</span>
                <div style="width: 15px; height: 15px; background-color: #ffac34; border: 1px solid #ccc; margin-right: 8px;"></div>
                <span style="margin-right: 15px; font-size: 14px;">Documentação não aprovada</span>
            </div>
            """,
            unsafe_allow_html=True
        )
       
    elif parametro == 'extrato':
        st.markdown(
            f"""
            <div style="display: flex; align-items: center; padding:10px; border:1px solid #ccc; border-radius:8px;">
                <div style="width: 15px; height: 15px; background-color: #e6937e; border: 1px solid #ccc; margin-right: 8px;"></div>
                <span style="margin-right: 15px; font-size: 14px;">Item do extrato sem despesa correspondente</span>
                <div style="width: 15px; height: 15px; background-color: #ffffae; border: 1px solid #ccc; margin-right: 8px;"></div>
                <span style="margin-right: 15px; font-size: 14px;">Mesma despesa mapeada com mais de um item no extrato bancário (correspondência incorreta)</span>
                <div style="width: 15px; height: 15px; background-color: #bfbfbf; border: 1px solid #ccc; margin-right: 8px;"></div>
                <span style="margin-right: 15px; font-size: 14px;">Despesas diferentes mapeadas com um mesmo item no extrato bancário (correspondência incorreta)</span>
            </div>
            """,
            unsafe_allow_html=True
        )

    
# Funções para formatar números
def format_brazilian(num):
    try:
        num = float(num)
        if abs(num) < 0.005:  # Trata -0.00 como 0.00
            num = 0.0
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return num

# 
def format_columns_brazilian(df, numeric_columns):
  for col in numeric_columns:
    if col in df.columns:
      df[col] = df[col].apply(format_brazilian)
  return df

# 
def format_percentage(num):
  try:
    num = float(num)
    formatted_num = f"{num * 100:,.2f}"  # Multiplica por 100 e formata
    return f"{formatted_num.replace(',', 'X').replace('.', ',').replace('X', '.')}%"  # Formata como percentual
  except (ValueError, TypeError):
    return num  # Retorna o valor original em caso de erro
  
# 
def format_columns_percentage(df, numeric_columns):
  for col in numeric_columns:
    if col in df.columns:
      df[col] = df[col].apply(format_percentage)
  return df  


# Função para formatar labels dos gráficos
def valores_labels_formatados(lista_valores):
    # Labels formatados
    labels = [format_brazilian(v) for v in lista_valores]

    # Dados com labels
    lista_valores_formatados = [
        {
            "value": v, 
            "label": {
                "show": True,
                "position": "bottom" if v < 0 else "top",  
                "formatter": lbl
            }
        }
        for v, lbl in zip(lista_valores, labels)
    ]
    return lista_valores_formatados


def _normalize_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower().strip()               # remove espaços no começo/fim
    text = re.sub(r"[-–—]", "-", text)        # normaliza todos os tipos de traço para "-"
    text = re.sub(r"\s+", " ", text)          # remove múltiplos espaços
    return text


def merge_com_fuzzy(df_custos, df_extratos, left_on, right_on, principal,
    text_left='Fornecedor', text_right='Descricao_Transacao', exceptions=exceptions, limiar=40):
    """
    Faz merge entre df_custos e df_extratos,
    e compara Fornecedor com Descricao_Transacao via fuzzy matching.
    exceptions = dict com pares manuais
    limiar = pontuação mínima de similaridade (0-100)
    """
    
    # merge 
    df_tmp = df_custos.merge(
        df_extratos, 
        left_on=left_on,
        right_on=right_on,
        how='left',
        suffixes=('_despesa','_extrato')
    )

    def calc_sim(row):
        f = _normalize_text(row[text_left])
        d = _normalize_text(row[text_right])

        # Exceções manuais
        for k, vs in exceptions.items():
            k_norm = _normalize_text(k)
            for v in (vs if isinstance(vs, list) else [vs]):
                v_norm = _normalize_text(v)
                if k_norm in f and v_norm in d:
                    return 100
                if v_norm in f and k_norm in d:
                    return 100
                # Regra: se a despesa não tiver descrição, aceita só Data + Valor
                if (f != "" and d == ""):
                    return 100   

        # Fuzzy padrão
        return fuzz.token_set_ratio(f, d)

    df_tmp['similaridade'] = df_tmp.apply(calc_sim, axis=1)

    # só mantém merge se atingir limiar
    extrato_cols = [c for c in df_extratos.columns if c not in df_custos.columns]
    df_tmp.loc[df_tmp['similaridade'] < limiar, extrato_cols] = None

    if principal == 'despesa':
        # mantém apenas o melhor match por despesa
        if 'ID_Despesa' in df_tmp.columns:
            df_tmp = (
                df_tmp.sort_values(by='similaridade', ascending=False)
                    .drop_duplicates(subset=['ID_Despesa'], keep='first')
            )

        # remove duplicatas de despesas que não encontraram correspondência no extrato
        mask_sem_extrato = df_tmp['ID_Extrato_Bancario'].isna()
        if 'ID_Despesa' in df_tmp.columns:  
            df_tmp_sem_extrato = (
                df_tmp[mask_sem_extrato]
                .drop_duplicates(subset=['ID_Despesa'])
            )
            df_tmp = pd.concat([
                df_tmp[~mask_sem_extrato],
                df_tmp_sem_extrato
            ], ignore_index=True)
    
    elif principal == 'extrato':
       # mantém apenas o melhor match por extrato
        if 'ID_Extrato_Bancario' in df_tmp.columns:
            df_tmp = (
                df_tmp.sort_values(by='similaridade', ascending=False)
                    .drop_duplicates(subset=['ID_Extrato_Bancario'], keep='first')
            )

        # remove duplicatas do extrato que não encontraram correspondência com alguma despesa
        # mask_sem_extrato = df_tmp['ID_Despesa'].isna()
        # if 'ID_Extrato_Bancario' in df_tmp.columns:  
        #     df_tmp_sem_extrato = (
        #         df_tmp[mask_sem_extrato]
        #         .drop_duplicates(subset=['ID_Extrato_Bancario'])
        #     )
        #     df_tmp = pd.concat([
        #         df_tmp[~mask_sem_extrato],
        #         df_tmp_sem_extrato
        #     ], ignore_index=True)

    return df_tmp


# Componente: botão exportar df em excel
def button_download(df, atributo, file_name, key):
    df_copia = df.copy()
    if atributo != 'ID_Despesa':
        df_copia["Despesa_Consta_Extrato"] = np.where(df_copia[atributo].isna(), "Não", "Sim")
    else:
        df_copia["Despesa_Correspondente"] = np.where(df_copia[atributo].isna(), "Não", "Sim")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_copia.to_excel(writer, index=False, sheet_name=file_name)
    excel_data = output.getvalue()

    st.download_button(
        label="Baixar Excel",
        data=excel_data,
        file_name=f"{file_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        icon=":material/download:",
        key=key
    )


# Funções excel
def export_to_excel(df, sheet_name, excel_filename):
  if os.path.exists(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
  else:
    wb = openpyxl.Workbook()

  if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])
  ws = wb.create_sheet(title=sheet_name)
  
  # Escrever os cabeçalhos
  for col_idx, column_title in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=column_title)
  
  # Escrever os dados
  for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
      ws.cell(row=row_idx, column=col_idx, value=value)

  wb.save(excel_filename)


# Funções - Fluxo de Caixa Futuro
def component_plotDataframe_aggrid(
    df: pd.DataFrame,
    name: str,
    num_columns: list = [],
    percent_columns: list = [],
    df_details: pd.DataFrame = None,
    coluns_merge_details: list = None,
    coluns_name_details: str = None,
    key: str = "default"
):
    """
    Exibe um DataFrame no AG Grid com:
    - filtros de conjunto em todas as colunas (agSetColumnFilter)
    - colunas numéricas com ordenação e filtro numérico (agNumberColumnFilter)
    - formatação brasileira de números e percentuais
    - colunas redimensionáveis dinamicamente para não cortar o cabeçalho
    """
    st.markdown(
        f"<h5 style='text-align: center; background-color: #ffb131; padding: 0.1em;'>{name}</h5>",
        unsafe_allow_html=True
    )

    # Formatter para números BR
    fmt_brl = JsCode(
        "function(params) { return params.value != null ? params.value.toLocaleString('pt-BR', { minimumFractionDigits: 2 }) : ''; }"
    )
    # Formatter para percentuais
    fmt_pct = JsCode(
        "function(params) { return params.value != null ? (params.value*100).toFixed(2).replace('.',',') + '%' : ''; }"
    )

    # Cria builder
    gb = GridOptionsBuilder.from_dataframe(df)
    # Configurações padrão para todas as colunas
    gb.configure_default_column(
        sortable=True,
        filter="agSetColumnFilter",
        resizable=True,
        editable=False,
        flex=1,
        minWidth=120,
        wrapHeaderText=True,
        autoHeaderHeight=True
    )

    # Converte e configura colunas numéricas
    for col in num_columns:
        if col in df.columns:
            series = df[col]
            # converte para float
            if not is_numeric_dtype(series):
                cleaned = series.astype(str).str.replace(r'[^\d,.-]', '', regex=True)
                cleaned = cleaned.str.replace(r'\.(?=\d{3},)', '', regex=True)
                numeric = cleaned.str.replace(',', '.', regex=False)
                df[col] = pd.to_numeric(numeric, errors='coerce')
            else:
                df[col] = series.astype(float)
            # configura coluna como numérica
            gb.configure_column(
                field=col,
                type=["numericColumn"],
                filter="agNumberColumnFilter",
                valueFormatter=fmt_brl
            )

    # Converte e configura colunas percentuais
    for col in percent_columns:
        if col in df.columns:
            series = df[col]
            if not is_numeric_dtype(series):
                cleaned = (
                    series.astype(str)
                          .str.replace('%', '', regex=False)
                          .str.replace(',', '.', regex=False)
                          .str.replace(r'[^\d.-]', '', regex=True)
                )
                df[col] = pd.to_numeric(cleaned, errors='coerce') / 100.0
            else:
                df[col] = series.astype(float)
            gb.configure_column(
                field=col,
                type=["numericColumn"],
                filter="agNumberColumnFilter",
                valueFormatter=fmt_pct
            )

    # Build options
    grid_options = gb.build()
    # Master-detail (opcional)
    if df_details is not None and coluns_merge_details and coluns_name_details:
        df['detail'] = df[coluns_merge_details].apply(
            lambda keys: df_details[df_details[coluns_merge_details] == keys].to_dict('records')
        )
        grid_options.update({
            "masterDetail": True,
            "columnDefs": [{"field": coluns_name_details, "cellRenderer": "agGroupCellRenderer"}] +
                           [{"field": c} for c in df.columns if c not in [coluns_name_details, 'detail']],
            "detailCellRendererParams": {
                "detailGridOptions": {"columnDefs": [{"field": c} for c in df_details.columns]},
                "getDetailRowData": JsCode("function(params){params.successCallback(params.data.detail);}")
            }
        })

    # Exibe grid
    df_to_show = df.copy()
    if 'detail' in df_to_show:
        df_to_show = df_to_show.drop(columns=['detail'])
    grid_response = AgGrid(
        df_to_show,
        gridOptions=grid_options,
        enable_enterprise_modules=True,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        fit_columns_on_grid_load=True,
        allow_unsafe_jscode=True,
        key=f"aggrid_{name}_{key}",
        theme=StAggridTheme(base="balham").withParams().withParts('colorSchemeDark')
    )
    return pd.DataFrame(grid_response['data'])

    

def function_copy_dataframe_as_tsv(df: pd.DataFrame):
    # Cria uma cópia do DataFrame para não modificar o original
    df_copy = df.copy()
    
    # Identifica colunas numéricas
    numeric_columns = df_copy.select_dtypes(include=[np.number]).columns
    
    # Formata colunas numéricas substituindo "." por ","
    for col in numeric_columns:
        df_copy[col] = df_copy[col].astype(str).str.replace('.', ',', regex=False)
    
    # Gera a string TSV
    tsv = df_copy.to_csv(index=False, sep='\t')

    # HTML+JS para ocultar o textarea, copiar ao clicar e avisar o usuário
    components.html(f"""
    <textarea id="clipboard-textarea" style="position: absolute; left: -10000px;">
{tsv}
    </textarea>
    <button onclick="
        const ta = document.getElementById('clipboard-textarea');
        ta.select();
        document.execCommand('copy');
        alert('TSV copiado para a área de transferência!');
    " style="
        background-color:#1e1e1e;
        color:#fff;
        border:1px solid #333;
        padding:8px 16px;
        border-radius:4px;
        cursor:pointer;
        margin-top:8px;
    ">Copiar como TSV</button>
    """, height=80)


    