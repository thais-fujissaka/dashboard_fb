import streamlit as st
import pandas as pd
import numpy as np
from workalendar.america import Brazil
import openpyxl
import os
from utils.functions.date_functions import *
from utils.user import *
import mysql.connector
from streamlit.logger import get_logger


LOGGER = get_logger(__name__)

def mysql_connection_fb():
	mysql_config = st.secrets["mysql_fb"]

	conn_fb = mysql.connector.connect(
			host=mysql_config['host'],
			port=mysql_config['port'],
			database=mysql_config['database'],
			user=mysql_config['username'],
			password=mysql_config['password']
		)    
	return conn_fb

def mysql_connection_eshows():
    mysql_config = st.secrets["mysql_eshows"]
    # Create MySQL connection
    conn = mysql.connector.connect(
        host=mysql_config['host'],
        port=mysql_config['port'],
        database=mysql_config['database'],
        user=mysql_config['username'],
        password=mysql_config['password']
    )    
    return conn


def execute_query(query, use_eshows=False):
    conn = (
        mysql_connection_fb() if use_eshows == False 
        else mysql_connection_eshows()
    )
    cursor = conn.cursor()
    try:
        cursor.execute(query)
        
        # Verifique se cursor.description não é None
        if cursor.description is None:
            print("Descrição do cursor é None")
            return None, None

        # Obter nomes das colunas
        column_names = [col[0] for col in cursor.description]
        # Obter resultados
        result = cursor.fetchall()
        
        if not result:
            print("Nenhuma linha retornada pela consulta.")
        
        cursor.close()
        conn.close()
        return result, column_names
    except Exception as e:
        print(f"Erro ao executar a consulta: {e}")
        return None, None
    finally:
        cursor.close()
        conn.close()


def dataframe_query(query, use_eshows=False):
	resultado, nomeColunas = execute_query(query, use_eshows=use_eshows)
	dataframe = pd.DataFrame(resultado, columns=nomeColunas)
	return dataframe


# Permissões de usuário
@st.cache_data
def GET_PERMISSIONS(login):
	loginStr = f"'{login}'"
	return dataframe_query(f''' 
		SELECT
            tcd.NOME_CARGO AS 'Permissao'
        FROM
            T_USUARIO_CARGO_DASH tucd 
            LEFT JOIN ADMIN_USERS au ON tucd.FK_USUARIO = au.ID 
            LEFT JOIN T_CARGO_DASH tcd ON tucd.FK_CARGO = tcd.ID 
        WHERE au.LOGIN = {loginStr}
  	''')


@st.cache_data
def GET_USERNAME(email):
	emailStr = f"'{email}'"
	return dataframe_query(f'''
		SELECT 
			au.FULL_NAME AS 'Nome'
		FROM
			ADMIN_USERS au 
		WHERE au.LOGIN = {emailStr}
  ''')


def config_permissoes_user():
    email = st.session_state.get("user_email", "Usuário desconhecido")
    dfpermissao = GET_PERMISSIONS(email)
    if dfpermissao.empty: # Não está no EPM
        permissao = ["Gazit"]
        nomeUser = ""
    else: # Está no EPM
        permissao = dfpermissao["Permissao"].tolist()
        nomeUser = GET_USERNAME(email)
        nomeUser = " ".join(nomeUser["Nome"].tolist())
    return permissao, nomeUser, email

def GET_ABAS_CARGOS(cargo):
    cargoStr = f"'{cargo}'"
    return dataframe_query(f'''
        SELECT
            tad.ID AS 'ID Aba',
            tad.NOME_ABA AS 'Aba'
        FROM
            T_CARGO_ABA_DASH tcad 
            LEFT JOIN T_CARGO_DASH tcd ON tcad.FK_CARGO = tcd.ID
            LEFT JOIN T_ABAS_DASH tad ON tcad.FK_ABA = tad.ID
        WHERE tcd.NOME_CARGO = {cargoStr}
    ''')

def GET_LOJAS():
  return dataframe_query(f'''
    SELECT
    te.ID as 'ID_Loja',
    te.NOME_FANTASIA as 'Loja'
    FROM T_EMPRESAS te
''')


@st.cache_data
def GET_LOJAS_USER(email):
	emailStr = f"'{email}'"
	return dataframe_query(f'''
		SELECT 
			te.NOME_FANTASIA AS 'Loja'
		FROM
			ADMIN_USERS au 
			LEFT JOIN T_USUARIOS_EMPRESAS tue ON au.ID = tue.FK_USUARIO 
			LEFT JOIN T_EMPRESAS te ON tue.FK_EMPRESA = te.ID
			LEFT JOIN T_LOJAS tl ON te.ID = tl.ID
		WHERE au.LOGIN = {emailStr}
  	''')

def config_sidebar():

    abas_secoes = {
        100: "KPI's de Faturamento",
        101: "KPI's de Faturamento",
        102: "KPI's de Faturamento",
        103: "KPI's de Faturamento",
        104: "KPI's de Faturamento",
        105: "KPI's de Faturamento",
        106: "KPI's de Faturamento - Eventos",
        107: "KPI's de Faturamento - Eventos",
        108: "KPI's de Faturamento - Eventos",
        109: "KPI's de Faturamento - Eventos",
        110: "KPI's de Faturamento - Eventos",
        111: "KPI's de Faturamento - Eventos",
        112: "KPI's de Faturamento - Eventos",
        113: "KPI's de Faturamento - Eventos",
        114: "KPI's de Faturamento - Eventos",
        115: "KPI's de Faturamento - Eventos",
        116: "KPI's de Faturamento - Eventos",
        117: "KPI's de Faturamento - Eventos",
        118: "Fluxo de Caixa",
        119: "Fluxo de Caixa",
        120: "Fluxo de Caixa",
        121: "Conciliação",
        122: "Conciliação",
        123: "Conciliação",
        124: "KPI's de Resultado Operacional",
        125: "KPI's de Resultado Operacional",
        126: "KPI's de Resultado Operacional",
        127: "KPI's de Resultado Operacional",
        128: "KPI's de Resultado Operacional",
        129: "KPI's de Resultado Operacional",
        130: "KPI's de Resultado Operacional - Suprimentos",
        131: "KPI's de Resultado Operacional - Suprimentos",
        132: "KPI's de Resultado Operacional - Suprimentos",
        133: "Auditoria",
    }

    cargo, user_name, email = config_permissoes_user()
    
    st.sidebar.header(f"Bem-vindo(a) {user_name}!")
    col1, col2 = st.sidebar.columns(2)
    with col1:
        st.button(label="Atualizar", on_click = st.cache_data.clear, width='stretch', icon='🔄')
    with col2:
        if st.button("Logout", width='stretch', icon='🚪'):
            logout()

    if st.session_state["loggedIn"]:
        abas_permitidas = st.session_state["abas_permitidas"]

        # Organizar abas por secao
        abas_por_secao = {}
        for aba in abas_permitidas:
            secao = abas_secoes.get(aba['ID Aba'])
            if secao:
                if secao not in abas_por_secao:
                    abas_por_secao[secao] = []
                abas_por_secao[secao].append(aba)
        
        # Renderizar cada secao
        for secao in abas_por_secao:
            st.sidebar.markdown(f"## {secao}")
            for aba in abas_por_secao[secao]:
                st.sidebar.page_link(f'{aba["page_link"]}', label=f'{aba["Aba"]}')
        
    else:
        st.sidebar.write("Por favor, faça login para acessar o menu.")


def filtrar_por_classe_selecionada(dataframe, classe, valores_selecionados):
    if valores_selecionados:
        dataframe = dataframe[dataframe[classe].isin(valores_selecionados)]
    return dataframe


def safe_sheet_name(name):
    # Remove caracteres inválidos e limita a 31 chars
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]


def export_to_excel(df, sheet_name, excel_filename):
    if os.path.exists(excel_filename):
        wb = openpyxl.load_workbook(excel_filename)
    else:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name)

    # Escrever os cabeçalhos
    for col_idx, column_title in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=column_title)

    # Escrever os dados
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(excel_filename)


def format_brazilian(num):
    try:
        num = float(num)
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return num
    

def format_brazilian_without_decimal(num):
    try:
        num = float(num)
        return f"{num:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return num


def format_columns_brazilian(df, numeric_columns):
    if df is not None and not df.empty:
        df = df.copy()
        for col in numeric_columns:
            if col in df.columns:
                df[col] = df[col].apply(format_brazilian)
        return df
    else:
        return df


def format_percentage(num):
    try:
        num = float(num)
        formatted_num = f"{num * 100:,.2f}"  # Multiplica por 100 e formata
        return f"{formatted_num.replace(',', 'X').replace('.', ',').replace('X', '.')}%"  # Formata como percentual
    except (ValueError, TypeError):
        return num  # Retorna o valor original em caso de erro


def format_columns_percentage(df, numeric_columns):
    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].apply(format_percentage)
    return df


# Dataframe filtrado pela casa:
def df_filtrar_casa(df, id_casa):
    df_filtrado = df[df["Casa"] == id_casa]
    return df_filtrado


def df_filtrar_periodo_data(df, coluna_data, data_inicio, data_fim):

    data_inicio = pd.to_datetime(data_inicio)
    data_fim = pd.to_datetime(data_fim) + pd.DateOffset(days=1)

    df = df.copy()

    df[coluna_data] = pd.to_datetime(df[coluna_data])
    df_filtrado = df.loc[
        (df[coluna_data] >= data_inicio) & (df[coluna_data] < data_fim)
    ]

    return df_filtrado


def df_filtrar_mes(df, coluna_data, mes):

    df = df.copy()

    df[coluna_data] = pd.to_datetime(df[coluna_data])

    df_filtrado = df.loc[(df[coluna_data].dt.month == int(mes))]

    return df_filtrado


def df_filtrar_ano(df, coluna_data, ano):

    df = df.copy()

    df[coluna_data] = pd.to_datetime(df[coluna_data])

    df_filtrado = df.loc[(df[coluna_data].dt.year == int(ano))]

    return df_filtrado


def escape_dolar(texto):
    return texto.replace('$', r'\$')


def highlight_values(val):
    color = 'red' if '-' in val else 'green'
    return f'color: {color}'


def preparar_dados_lojas_user_financeiro():
    permissao, nomeuser, username = config_permissoes_user()
    if 'Administrador' in permissao:
        dflojas = GET_LOJAS()
        lojasARemover = ['Casa Teste', 'Casa Teste 2', 'Casa Teste 3']
        dflojas = dflojas[~dflojas['Loja'].isin(lojasARemover)]
    else:
        dflojas = GET_LOJAS_USER(username)

    lojasReais = [
        'Abaru - Priceless', 'Arcos', 'Bar Brahma - Centro', 'Bar Brahma Paulista', 'Bar Léo - Centro',
        'Blue Note - São Paulo', 'Blue Note SP (Novo)', 'Delivery Bar Leo Centro', 'Delivery Fabrica de Bares',
        'Delivery Jacaré', 'Delivery Orfeu', 'Edificio Rolim', 'Escritório Fabrica de Bares',
        'Girondino ', 'Girondino - CCBB', 'Hotel Maraba', 'Jacaré', 'Love Cabaret', 'Notiê - Priceless',
        'Orfeu', 'Priceless', 'Riviera Bar', 'Sanduiche comunicação LTDA ', 'Tempus Fugit  Ltda ',
        'Ultra Evil Premium Ltda ', 'Bar Brahma - Granja', 'Brahma - Ribeirão', 'The Cavern'
    ]

    lojas = dflojas[dflojas['Loja'].isin(set(lojasReais))]['Loja'].tolist()
    lojas.sort(key=str.lower)

    # Verificar se ambas as lojas estão na lista
    if 'Abaru - Priceless' in lojas and 'Notiê - Priceless' in lojas:
        # Remover a 'loja 1' da lista
        lojas.remove('Abaru - Priceless')

        # Encontrar o índice da 'loja 3' para inserir a 'loja 1' logo após
        indice_loja_alvo = lojas.index('Notiê - Priceless')

        # Inserir a 'loja 1' após a 'loja 3'
        lojas.insert(indice_loja_alvo + 1, 'Abaru - Priceless')

    return lojas


def preparar_dados_lojas_user_projecao_fluxo():
  permissao, nomeuser, username = config_permissoes_user()
  if 'Administrador' or 'Acesso Financeiro 3' in permissao:
    dflojas = GET_LOJAS()
    lojasARemover = ['Casa Teste', 'Casa Teste 2', 'Casa Teste 3']
    dflojas = dflojas[~dflojas['Loja'].isin(lojasARemover)]
  else:
    dflojas = GET_LOJAS_USER(username)

  lojasReais = ['Abaru - Priceless', 'Arcos', 'All bar', 'Bar Brahma Aeroclube', 'Brahma Aricanduva',
                'Bar Brahma - Centro', 'Bar Brahma Paulista', 'Bar Brasilia -  Aeroporto', 'Bardassê', 'Bar Léo - Centro', 'Bar Léo - Vila Madalena', 'Blue Note - São Paulo', 'Blue Note SP (Novo)',
                'Colorado Aeroporto BSB', 'Delivery Bar Leo Centro', 'Delivery Fabrica de Bares', 'Delivery Jacaré', 'Delivery Orfeu', 'Duroc ', 'Edificio Rolim', 'Escritório Fabrica de Bares', 'FDB DIGITAL PARTICIPACOES LTDA', 'FDB HOLDING INFERIOR LTDA', 'FDB HOLDING SUPERIOR LTDA', 'Filial', 'Hbar participacoes e empreendimentos ', 'Ilha das Flores ', 'Lojinha - Brahma', 'Navarro', 'Patizal ',  'Piratininga', 'Tundra',
                'Girondino ', 'Girondino - CCBB', 'Hotel Maraba', 'Jacaré', 'Love Cabaret', 'Notiê - Priceless', 'Orfeu', 'Priceless', 'Riviera Bar', 
                'Sanduiche comunicação LTDA ', 'Tempus Fugit  Ltda ', 'The Cavern', 'Ultra Evil Premium Ltda ', 'Bar Brahma - Granja', 'Brahma - Ribeirão']
  
  lojas = dflojas[dflojas['Loja'].isin(set(lojasReais))]['Loja'].tolist()
  lojas.sort(key=str.lower)

  # Verificar se ambas as lojas estão na lista
  if 'Abaru - Priceless' in lojas and 'Notiê - Priceless' in lojas:
    # Remover a 'loja 1' da lista
    lojas.remove('Abaru - Priceless')
    
    # Encontrar o índice da 'loja 3' para inserir a 'loja 1' logo após
    indice_loja_alvo = lojas.index('Notiê - Priceless')
    
    # Inserir a 'loja 1' após a 'loja 3'
    lojas.insert(indice_loja_alvo + 1, 'Abaru - Priceless')

    return lojas


def obter_valores_unicos_ordenados(df, coluna):
    dados = df[coluna].dropna().unique().tolist()
    dados.sort(key=str.lower)
    return dados


def highlight_values_inverse(val):
  if '-' in val:
    color = 'green' 
  elif val == '0,00' or val == 'nan':
    color = 'black'
  else:
    color = 'red'
  return f'color: {color}'



